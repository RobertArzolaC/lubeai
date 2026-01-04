import openpyxl
from django.contrib import messages
from django.contrib.auth.mixins import (
    LoginRequiredMixin,
    PermissionRequiredMixin,
)
from django.contrib.messages.views import SuccessMessageMixin
from django.db import transaction
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.views.generic import (
    CreateView,
    DetailView,
    FormView,
    UpdateView,
    View,
)
from django_filters.views import FilterView
from openpyxl.styles import Font

from apps.core import mixins as core_mixins
from apps.equipment import filtersets, forms, models
from apps.users.models import Organization


class MachineListView(
    PermissionRequiredMixin, FilterView, LoginRequiredMixin, SuccessMessageMixin
):
    model = models.Machine
    permission_required = "equipment.view_machine"
    filterset_class = filtersets.MachineFilter
    template_name = "equipment/machine/list.html"
    context_object_name = "machines"
    paginate_by = 5

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["entity"] = _("Machine")
        context["entity_plural"] = _("Machines")
        context["back_url"] = reverse_lazy("apps.dashboard:index")
        context["add_entity_url"] = reverse_lazy(
            "apps.equipment:machine_create"
        )
        return context


class MachineDetailView(
    PermissionRequiredMixin, DetailView, LoginRequiredMixin
):
    model = models.Machine
    permission_required = "equipment.view_machine"
    template_name = "equipment/machine/detail.html"
    context_object_name = "machine"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["entity"] = _("Machine")
        context["back_url"] = reverse_lazy("apps.equipment:machine_list")
        context["edit_url"] = reverse_lazy(
            "apps.equipment:machine_update", kwargs={"pk": self.object.pk}
        )
        context["components"] = self.object.components.select_related("type")
        return context


class MachineCreateView(
    PermissionRequiredMixin, CreateView, LoginRequiredMixin, SuccessMessageMixin
):
    model = models.Machine
    form_class = forms.MachineForm
    permission_required = "equipment.add_machine"
    template_name = "equipment/machine/form.html"
    success_message = _("Machine created successfully")
    success_url = reverse_lazy("apps.equipment:machine_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["entity"] = _("Machine")
        context["back_url"] = reverse_lazy("apps.equipment:machine_list")
        if self.request.POST:
            context["component_formset"] = forms.ComponentFormSet(
                self.request.POST
            )
        else:
            context["component_formset"] = forms.ComponentFormSet()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        component_formset = context["component_formset"]
        if component_formset.is_valid():
            self.object = form.save()
            component_formset.instance = self.object
            component_formset.save()
            return super().form_valid(form)
        return self.render_to_response(context)


class MachineUpdateView(
    PermissionRequiredMixin, UpdateView, LoginRequiredMixin, SuccessMessageMixin
):
    model = models.Machine
    form_class = forms.MachineForm
    permission_required = "equipment.change_machine"
    template_name = "equipment/machine/form.html"
    success_message = _("Machine updated successfully")
    success_url = reverse_lazy("apps.equipment:machine_list")
    context_object_name = "machine"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["entity"] = _("Machine")
        context["back_url"] = reverse_lazy("apps.equipment:machine_list")
        if self.request.POST:
            context["component_formset"] = forms.ComponentFormSet(
                self.request.POST, instance=self.object
            )
        else:
            context["component_formset"] = forms.ComponentFormSet(
                instance=self.object
            )
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        component_formset = context["component_formset"]
        if component_formset.is_valid():
            self.object = form.save()
            component_formset.instance = self.object
            component_formset.save()
            return super().form_valid(form)
        return self.render_to_response(context)


class MachineDeleteView(core_mixins.AjaxDeleteViewMixin):
    model = models.Machine


class MachineBulkUploadView(
    PermissionRequiredMixin, LoginRequiredMixin, SuccessMessageMixin, FormView
):
    """View for bulk uploading machines from Excel file."""

    permission_required = "equipment.add_machine"
    form_class = forms.MachineBulkUploadForm
    template_name = "equipment/machine/bulk_upload.html"
    success_url = reverse_lazy("apps.equipment:machine_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["entity"] = _("Machine")
        context["back_url"] = reverse_lazy("apps.equipment:machine_list")
        return context

    def form_valid(self, form):
        """Process the uploaded Excel file."""
        excel_file = form.cleaned_data["file"]
        results = self._process_excel_file(excel_file)

        # Display results to user
        if results["created"]:
            messages.success(
                self.request,
                _(f"Successfully created {results['created']} machines"),
            )
        if results["updated"]:
            messages.info(
                self.request,
                _(f"Successfully updated {results['updated']} machines"),
            )
        if results["skipped"]:
            messages.info(
                self.request,
                _(f"Skipped {results['skipped']} header/title rows"),
            )
        if results["errors"]:
            for error in results["errors"]:
                messages.error(self.request, error)

        return super().form_valid(form)

    def _process_excel_file(self, excel_file):
        """Process the Excel file and create/update machines."""
        results = {"created": 0, "updated": 0, "errors": [], "skipped": 0}

        try:
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active

            # Skip header rows (row 1: title, row 2: column headers)
            for row_num, row in enumerate(
                worksheet.iter_rows(min_row=3, values_only=True), start=3
            ):
                if not any(row):  # Skip empty rows
                    continue

                # Skip repeated title rows and header rows
                if self._should_skip_row(row):
                    results["skipped"] += 1
                    continue

                try:
                    with transaction.atomic():
                        result = self._process_row(row, row_num)
                        if result["action"] == "created":
                            results["created"] += 1
                        elif result["action"] == "updated":
                            results["updated"] += 1
                except Exception as e:
                    results["errors"].append(_(f"Row {row_num}: {str(e)}"))

        except Exception as e:
            results["errors"].append(_(f"Error reading Excel file: {str(e)}"))

        return results

    def _should_skip_row(self, row):
        """
        Check if a row should be skipped (titles, headers, etc).

        Args:
            row: Tuple of cell values from the Excel row

        Returns:
            bool: True if the row should be skipped
        """
        # Convert all values to string for comparison
        row_values = [str(cell).strip().upper() if cell else "" for cell in row]

        # Skip if first cell contains title text
        if row_values[0] and "REPORTE DE EQUIPOS" in row_values[0]:
            return True

        # Skip if it's a header row (check for common header patterns)
        header_patterns = [
            "NOMBRE DE EQUIPO",
            "CLIENTE",
            "DESCRIPCIÓN",
            "MODELO",
            "COMPONENTES ASOCIADOS",
            "ESTADO",
        ]

        # If first cell matches any header pattern, skip the row
        if any(pattern in row_values[0] for pattern in header_patterns):
            return True

        # Skip if multiple columns contain header patterns (header row detection)
        header_matches = sum(
            1
            for cell in row_values
            if any(pattern in cell for pattern in header_patterns)
        )
        if header_matches >= 2:  # If 2 or more columns contain header text
            return True

        return False

    def _process_row(self, row, row_num):
        """Process a single row from the Excel file."""
        # Map columns: A→name, B→organization, C→serial_number, D→model, E→components
        name = row[0].strip() if row[0] else ""
        organization_name = row[1].strip() if row[1] else ""
        serial_number = row[2].strip() if row[2] else ""
        model = row[3].strip() if row[3] else ""
        components = row[4].strip() if len(row) > 4 and row[4] else ""

        # Validate required fields
        if not name:
            raise ValueError(_("Machine name is required"))
        if not serial_number:
            raise ValueError(_("Serial number is required"))
        if not model:
            raise ValueError(_("Model is required"))

        # Find organization if provided
        organization = None
        if organization_name:
            try:
                organization = Organization.objects.get(
                    name=organization_name, is_active=True
                )
            except Organization.DoesNotExist:
                raise ValueError(
                    _(f"Organization '{organization_name}' not found")
                )

        # Create or update machine
        machine, created = models.Machine.objects.update_or_create(
            serial_number=serial_number,
            defaults={
                "name": name,
                "model": model,
                "organization": organization,
                "is_active": True,
                "created_by": self.request.user,
                "modified_by": self.request.user,
            },
        )

        # Process components if provided
        if components:
            self._process_components(machine, components)

        return {
            "action": "created" if created else "updated",
            "machine": machine,
        }

    def _process_components(self, machine, components_str):
        """Process components for a machine."""
        if not components_str:
            return

        component_names = [name.strip() for name in components_str.split(",")]

        for component_name in component_names:
            if not component_name:
                continue

            # Get or create component type
            component_type, created = (
                models.ComponentType.objects.get_or_create(
                    name=component_name,
                    defaults={
                        "description": "Auto-created from bulk upload",
                        "is_active": True,
                        "created_by": self.request.user,
                        "modified_by": self.request.user,
                    },
                )
            )

            # Create component with same serial number as machine
            models.Component.objects.update_or_create(
                machine=machine,
                type=component_type,
                defaults={
                    "serial_number": machine.serial_number,
                    "installation_datetime": timezone.now(),
                    "is_active": True,
                    "created_by": self.request.user,
                    "modified_by": self.request.user,
                },
            )


class MachineBulkTemplateView(
    PermissionRequiredMixin, LoginRequiredMixin, View
):
    """View to download Excel template for bulk upload."""

    permission_required = "equipment.add_machine"

    def get(self, request, *args, **kwargs):
        """Generate and return Excel template."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Machines"

        # Title row
        worksheet["A1"] = "REPORTE DE EQUIPOS"
        worksheet.merge_cells("A1:E1")
        title_cell = worksheet["A1"]
        title_cell.font = Font(size=14, bold=True)

        # Header row
        headers = [
            "NOMBRE DE EQUIPO",
            "CLIENTE",
            "NÚMERO DE SERIE",
            "MODELO",
            "COMPONENTES ASOCIADOS",
        ]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)

        # Example data row
        example_data = [
            "A5K-837 / NEVALLOS TUBILLAS PABLO",
            "NEUMA PERU",
            "A5K-837",
            "FREIGHTLINE R M2-212",
            "MOTOR",
        ]
        for col, data in enumerate(example_data, 1):
            worksheet.cell(row=3, column=col, value=data)

        # Adjust column widths
        column_widths = [40, 20, 15, 25, 30]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = width

        # Create response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="machine_bulk_template.xlsx"'
        )

        workbook.save(response)
        return response
