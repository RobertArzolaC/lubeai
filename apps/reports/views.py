import logging

import openpyxl
from django.contrib import messages
from django.contrib.auth.mixins import (
    LoginRequiredMixin,
    PermissionRequiredMixin,
)
from django.contrib.messages.views import SuccessMessageMixin
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy as _
from django.views.generic import (
    CreateView,
    DetailView,
    FormView,
    UpdateView,
    View,
)
from django_filters.views import FilterView
from openpyxl.styles import Font

from apps.core import mixins as core_mixins
from apps.reports import filtersets, forms, models
from apps.reports.services.bulk_upload import ReportBulkUploadService

logger = logging.getLogger(__name__)


class ReportListView(
    PermissionRequiredMixin,
    FilterView,
    LoginRequiredMixin,
):
    """List view for reports."""

    model = models.Report
    permission_required = "reports.view_report"
    filterset_class = filtersets.ReportFilter
    template_name = "reports/report/list.html"
    context_object_name = "reports"
    paginate_by = 5

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["entity"] = _("Report")
        context["entity_plural"] = _("Reports")
        context["back_url"] = reverse_lazy("apps.dashboard:index")
        context["add_entity_url"] = reverse_lazy("apps.reports:report_create")
        context["bulk_upload_url"] = reverse_lazy(
            "apps.reports:report_bulk_upload"
        )
        return context


class ReportDetailView(
    PermissionRequiredMixin,
    DetailView,
    LoginRequiredMixin,
):
    """Detail view for report."""

    model = models.Report
    permission_required = "reports.view_report"
    template_name = "reports/report/detail.html"
    context_object_name = "report"

    def get_queryset(self):
        """Optimize query with select_related for analysis."""
        return super().get_queryset().select_related("analysis")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["entity"] = _("Report")
        context["back_url"] = reverse_lazy("apps.reports:report_list")
        context["edit_url"] = reverse_lazy(
            "apps.reports:report_update", kwargs={"pk": self.object.pk}
        )
        return context


class ReportCreateView(
    PermissionRequiredMixin,
    CreateView,
    LoginRequiredMixin,
    SuccessMessageMixin,
):
    """Create view for report."""

    model = models.Report
    form_class = forms.ReportForm
    permission_required = "reports.add_report"
    template_name = "reports/report/form.html"
    success_message = _("Report created successfully")
    success_url = reverse_lazy("apps.reports:report_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["entity"] = _("Report")
        context["back_url"] = reverse_lazy("apps.reports:report_list")
        context["form_title"] = _("Create Report")
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.modified_by = self.request.user
        return super().form_valid(form)


class ReportUpdateView(
    PermissionRequiredMixin,
    UpdateView,
    LoginRequiredMixin,
    SuccessMessageMixin,
):
    """Update view for report."""

    model = models.Report
    form_class = forms.ReportForm
    permission_required = "reports.change_report"
    template_name = "reports/report/form.html"
    success_message = _("Report updated successfully")
    success_url = reverse_lazy("apps.reports:report_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["entity"] = _("Report")
        context["back_url"] = reverse_lazy("apps.reports:report_list")
        context["form_title"] = _("Edit Report")
        return context

    def form_valid(self, form):
        form.instance.modified_by = self.request.user
        return super().form_valid(form)


class ReportDeleteView(core_mixins.AjaxDeleteViewMixin):
    """Delete view for report."""

    model = models.Report


class ReportBulkUploadView(
    PermissionRequiredMixin,
    LoginRequiredMixin,
    SuccessMessageMixin,
    FormView,
):
    """Bulk upload view for reports."""

    permission_required = "reports.add_report"
    form_class = forms.ReportBulkUploadForm
    template_name = "reports/report/bulk_upload.html"
    success_url = reverse_lazy("apps.reports:report_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["entity"] = _("Report")
        context["entity_plural"] = _("Reports")
        context["back_url"] = reverse_lazy("apps.reports:report_list")
        context["template_url"] = reverse_lazy(
            "apps.reports:report_bulk_template"
        )
        return context

    def form_valid(self, form):
        excel_file = form.cleaned_data["file"]
        user_email = self.request.user.email

        logger.info(
            f"Starting bulk upload - User: {user_email}, File: {excel_file.name}"
        )

        try:
            service = ReportBulkUploadService(self.request.user)
            results = service.process_file(excel_file)
            self._show_results_messages(results)
        except Exception as e:
            error_msg = str(e)
            logger.exception(
                f"Bulk upload file processing error - User: {user_email}, "
                f"File: {excel_file.name}, Error: {error_msg}"
            )
            messages.error(
                self.request,
                _("Error processing file: %(error)s") % {"error": error_msg},
            )
        return super().form_valid(form)

    def _show_results_messages(self, results):
        """Show result messages to user."""
        if results["created"]:
            messages.success(
                self.request,
                _("Successfully created %(count)d reports")
                % {"count": results["created"]},
            )

        if results["skipped"]:
            messages.warning(
                self.request,
                _(
                    "Skipped %(count)d duplicate reports (lab numbers already exist)"
                )
                % {"count": results["skipped"]},
            )

        if results["errors"]:
            for error in results["errors"]:
                # Handle both string and dict error formats
                if isinstance(error, dict):
                    error_msg = (
                        f"Row {error.get('row_number', 'N/A')}: "
                        f"{error.get('error', str(error))}"
                    )
                else:
                    error_msg = str(error)
                messages.error(self.request, error_msg)

        logger.info(
            f"Bulk upload summary displayed - User: {self.request.user.email}, "
            f"Created: {results['created']}, "
            f"Skipped: {results['skipped']}, Errors: {len(results['errors'])}"
        )


class ReportBulkTemplateView(
    PermissionRequiredMixin,
    LoginRequiredMixin,
    View,
):
    """Generate Excel template for bulk upload."""

    permission_required = "reports.add_report"

    def get(self, request, *args, **kwargs):
        """Generate and return Excel template."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Reports Template"

        # Add title
        worksheet.merge_cells("A1:BF1")
        worksheet["A1"] = "REPORTE DE INSPECCIONES - TEMPLATE"
        worksheet["A1"].font = Font(bold=True, size=14)

        # Add headers for all 57 columns
        headers = [
            "N°",
            "No. Lab *",
            "Cliente",
            "Equipo",
            "Componente",
            "Cód/Núm Serie",
            "Lubricante",
            "Fecha Muestra",
            "Equipo Horas/Kms",
            "Lubricante Horas/Kms",
            "Fecha de Recepción",
            "Fecha de Reporte",
            "Cambio de Filtro",
            "Cambio de Aceite",
            "No.Per",
            "Otros",
            "Condición",
            "Comentario",
            "ITS 009/18 - Agua (Crackle test)",
            "ASTM D 95-13 - Agua por destilación",
            "ASTM D 7279-20 - Viscosidad a 40°C",
            "ASTM D 7279-20 - Viscosidad a 100°C",
            "ILT-096 - COMPATIBILIDAD",
            "ASTM D 2896-21 - Número Básico (TBN)",
            "ASTM D 664-24 - Número Acido (TAN)",
            "ASTM E 2412-23 - Oxidación",
            "ASTM E 2412-23 - Hollín",
            "ASTM E 2412-23 - Nitración",
            "ASTM E 2412-23 - Sulfatación",
            "ASTM E 2412-23 - Glicol",
            "ASTM E 2412-23 - Dilución",
            "ASTM E 2412-23 - Agua FTIR",
            "ITS 044/15 - PQ Index",
            "ISO 4406:2021 - Conteo de Partículas",
            "ASTM D 5185-18 - Hierro (Fe)",
            "ASTM D 5185-18 - Cromo (Cr)",
            "ASTM D 5185-18 - Plomo (Pb)",
            "ASTM D 5185-18 - Cobre (Cu)",
            "ASTM D 5185-18 - Estaño (Sn)",
            "ASTM D 5185-18 - Aluminio (Al)",
            "ASTM D 5185-18 - Níquel (Ni)",
            "ASTM D 5185-18 - Plata (Ag)",
            "ASTM D 5185-18 - Silicio (Si)",
            "ASTM D 5185-18 - Boro (B)",
            "ASTM D 5185-18 - Sodio (Na)",
            "ASTM D 5185-18 - Magnesio (Mg)",
            "ASTM D 5185-18 - Molibdeno (Mo)",
            "ASTM D 5185-18 - Titanio (Ti)",
            "ASTM D 5185-18 - Vanadio (V)",
            "ASTM D 5185-18 - Manganeso (Mn)",
            "ASTM D 5185-18 - Potasio (K)",
            "ASTM D 5185-18 - Fósforo (P)",
            "ASTM D 5185-18 - Zinc (Zn)",
            "ASTM D 5185-18 - Calcio (Ca)",
            "ASTM D 5185-18 - Bario (Ba)",
            "ASTM D 5185-18 - Cadmio (Cd)",
            "Apariencia - Visual",
        ]

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)

        # Add example data (basic report data only)
        example_data = [
            "1",  # N°
            "20001L-25",  # No. Lab
            "NEUMA PERU",  # Cliente
            "TRANSPORTES SATURNO / BUO-805",  # Equipo
            "MOTOR",  # Componente
            "BUO-805",  # Cód/Núm Serie
            "PETRONAS URANIA 15W40 CK-4",  # Lubricante
            "30/11/2025",  # Fecha Muestra
            "10377",  # Equipo Horas/Kms
            "720",  # Lubricante Horas/Kms
            "18/12/2025",  # Fecha de Recepción
            "22/12/2025",  # Fecha de Reporte
            "-",  # Cambio de Filtro
            "-",  # Cambio de Aceite
            "15886-25",  # No.Per
            "",  # Otros
            "Normal",  # Condición
            "Example notes",  # Comentario
        ]
        # Add empty values for lab analysis columns (18-56)
        example_data.extend([""] * 39)

        for col, value in enumerate(example_data, 1):
            worksheet.cell(row=3, column=col, value=value)

        # Set column widths (basic width for all columns)
        for col in range(1, 58):  # 57 columns
            worksheet.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = 15

        # Generate response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="report_bulk_template.xlsx"'
        )

        workbook.save(response)
        workbook.close()
        return response
