from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import (
    Count,
    Q,
)
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.views.generic import TemplateView, View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.dashboard.filtersets import ReportFilter
from apps.equipment.models import Machine
from apps.reports.choices import ReportCondition, ReportStatus
from apps.reports.models import Report
from apps.users.mixins import OrganizationRequiredMixin
from apps.users.models import Organization


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = "dashboard/index.html"

    def is_admin_user(self):
        """Check if user is staff or superuser."""
        return self.request.user.is_staff or self.request.user.is_superuser

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now()

        # Check if user is admin (staff/superuser)
        if self.is_admin_user():
            # Administrative dashboard - show all reports
            reports_qs = Report.objects.filter(is_active=True)

            # Basic statistics
            total_reports = reports_qs.count()

            # Count by condition
            condition_stats_raw = dict(
                reports_qs.values_list("condition").annotate(count=Count("id"))
            )

            # Ensure all condition choices are represented with default 0 values
            condition_stats = {}
            condition_stats_parts = {}
            for choice, _ in ReportCondition.choices:
                translated_label = str(dict(ReportCondition.choices)[choice])
                condition_stats[choice] = condition_stats_raw.get(choice, 0)
                condition_stats_parts[translated_label] = (
                    condition_stats_raw.get(choice, 0)
                )

            # Count by status
            status_stats_raw = dict(
                reports_qs.values_list("status").annotate(count=Count("id"))
            )

            # Ensure all status choices are represented with default 0 values
            status_stats = {}
            for choice, _ in ReportStatus.choices:
                translated_label = str(dict(ReportStatus.choices)[choice])
                status_stats[translated_label] = status_stats_raw.get(choice, 0)

            # Historical data by month (all time)
            monthly_data = (
                reports_qs.filter(sample_date__isnull=False)
                .annotate(month=TruncMonth("sample_date"))
                .values("month")
                .annotate(count=Count("id"))
                .order_by("month")
            )

            # Convert month datetime objects to strings for JSON serialization
            monthly_data = [
                {
                    "month": item["month"].strftime("%Y-%m-%d"),
                    "count": item["count"],
                }
                for item in monthly_data
            ]

            # Get all organizations for filter
            organizations = Organization.objects.filter(
                is_active=True, is_removed=False
            )

            # Get all machines (will be filtered dynamically)
            machines = Machine.objects.filter(is_active=True)

            # Latest reports
            latest_reports = reports_qs.order_by("-sample_date", "-created")[
                :10
            ]

            context.update(
                {
                    "is_admin": True,
                    "total_reports": total_reports,
                    "condition_stats": condition_stats,
                    "condition_stats_parts": condition_stats_parts,
                    "status_stats": status_stats,
                    "monthly_data": list(monthly_data),
                    "organizations": organizations,
                    "machines": machines,
                    "latest_reports": latest_reports,
                    "condition_choices": ReportCondition.choices,
                    "status_choices": ReportStatus.choices,
                }
            )
        else:
            # Check if user has organization access
            user_profile = getattr(self.request.user, "account", None)
            has_organization = bool(
                user_profile
                and user_profile.organization
                and user_profile.organization.is_active
                and not user_profile.organization.is_removed
            )

            # If user has organization, provide dashboard data
            if has_organization:
                organization = user_profile.organization
                context.update(
                    {
                        "is_admin": False,
                        "has_organization": has_organization,
                        "organization": organization,
                    }
                )
            else:
                context.update(
                    {
                        "is_admin": False,
                        "has_organization": has_organization,
                    }
                )

        context.update(
            {
                "last_updated": today,
            }
        )

        return context


class DashboardDataAPIView(LoginRequiredMixin, View):
    """
    AJAX endpoint to get filtered dashboard data for charts and tables.
    Only accessible by staff/superuser.
    """

    def get(self, request, *args, **kwargs):
        # Only allow staff/superuser access
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({"error": "Admin access required"}, status=403)

        # Base queryset - all reports for admin
        reports_qs = Report.objects.filter(is_active=True)

        # Apply filters using DashboardReportFilter
        filter_data = {
            "organization": request.GET.get("organization_id"),
            "start_date": request.GET.get("start_date"),
            "end_date": request.GET.get("end_date"),
            "machine": request.GET.get("machine_id"),
            "condition": request.GET.get("condition"),
            "status": request.GET.get("status"),
        }

        # Remove None values
        filter_data = {k: v for k, v in filter_data.items() if v}

        # Apply filter
        filterset = ReportFilter(filter_data, queryset=reports_qs)
        reports_qs = filterset.qs

        # Get updated statistics
        total_reports = reports_qs.count()

        # Count by condition
        condition_data = list(
            reports_qs.values("condition")
            .annotate(count=Count("id"))
            .order_by("condition")
        )

        # Count by status
        status_data = list(
            reports_qs.values("status")
            .annotate(count=Count("id"))
            .order_by("status")
        )

        # Historical data by month
        monthly_data = list(
            reports_qs.filter(sample_date__isnull=False)
            .annotate(month=TruncMonth("sample_date"))
            .values("month")
            .annotate(count=Count("id"))
            .order_by("month")
        )

        # Latest reports (limited for performance)
        latest_reports_data = []
        latest_reports = reports_qs.order_by("-sample_date", "-created")[:10]

        for report in latest_reports:
            latest_reports_data.append(
                {
                    "id": report.id,
                    "lab_number": report.lab_number,
                    "machine_name": report.machine.name
                    if report.machine
                    else "",
                    "condition": report.get_condition_display(),
                    "status": report.get_status_display(),
                    "sample_date": report.sample_date.strftime("%Y-%m-%d")
                    if report.sample_date
                    else "",
                    "condition_class": self._get_condition_class(
                        report.condition
                    ),
                    "status_class": self._get_status_class(report.status),
                }
            )

        return JsonResponse(
            {
                "total_reports": total_reports,
                "condition_data": condition_data,
                "status_data": status_data,
                "monthly_data": monthly_data,
                "latest_reports": latest_reports_data,
            }
        )

    def _get_condition_class(self, condition):
        """Get CSS class for condition display."""
        condition_classes = {
            ReportCondition.NORMAL: "success",
            ReportCondition.CAUTION: "warning",
            ReportCondition.CRITICAL: "danger",
        }
        return condition_classes.get(condition, "secondary")

    def _get_status_class(self, status):
        """Get CSS class for status display."""
        status_classes = {
            ReportStatus.PENDING: "warning",
            ReportStatus.REVIEWED: "info",
            ReportStatus.APPROVED: "success",
            ReportStatus.REJECTED: "danger",
        }
        return status_classes.get(status, "secondary")


class MachinesByOrganizationAPIView(LoginRequiredMixin, View):
    """
    AJAX endpoint to get machines by organization for dynamic filtering.
    Only accessible by staff/superuser.
    """

    def get(self, request, *args, **kwargs):
        # Only allow staff/superuser access
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({"error": "Admin access required"}, status=403)

        organization_id = request.GET.get("organization_id")

        if organization_id:
            try:
                organization_id = int(organization_id)
                machines = Machine.objects.filter(
                    organization_id=organization_id, is_active=True
                ).values("id", "name")
            except (ValueError, TypeError):
                machines = []
        else:
            machines = Machine.objects.filter(is_active=True).values(
                "id", "name"
            )

        return JsonResponse({"machines": list(machines)})


class ExportPageView(
    LoginRequiredMixin, OrganizationRequiredMixin, TemplateView
):
    """
    Page for exporting dashboard data with filters and preview.
    """

    template_name = "dashboard/export.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        if self.has_organization_access():
            organization = self.get_user_organization()

            # Get machines for the organization
            machines = Machine.objects.filter(
                organization=organization, is_active=True
            )

            context.update(
                {
                    "machines": machines,
                    "condition_choices": ReportCondition.choices,
                    "status_choices": ReportStatus.choices,
                    "max_records": 10000,
                }
            )

        return context


class ExportPreviewAPIView(LoginRequiredMixin, OrganizationRequiredMixin, View):
    """
    AJAX endpoint to preview how many records will be exported.
    """

    def get(self, request, *args, **kwargs):
        if not self.has_organization_access():
            return JsonResponse({"error": "Organization required"}, status=403)

        organization = self.get_user_organization()

        # Base queryset
        reports_qs = Report.objects.filter(
            organization=organization, is_active=True
        )

        # Prepare filter data
        filter_data = {
            "start_date": request.GET.get("start_date"),
            "end_date": request.GET.get("end_date"),
            "machine": request.GET.get("machine_id"),
            "condition": request.GET.get("condition"),
            "status": request.GET.get("status"),
        }

        # Remove None values
        filter_data = {k: v for k, v in filter_data.items() if v}

        # Apply filters using ReportFilter
        filterset = ReportFilter(filter_data, queryset=reports_qs)
        reports_qs = filterset.qs

        # Count records
        total_records = reports_qs.count()
        will_export = min(total_records, 10000)

        return JsonResponse(
            {
                "total_records": total_records,
                "will_export": will_export,
                "has_limit": total_records > 10000,
                "limit": 10000,
            }
        )


class DashboardExportView(LoginRequiredMixin, OrganizationRequiredMixin, View):
    """
    Export filtered dashboard data to Excel format.
    Limited to 10,000 records for performance.
    """

    def get(self, request, *args, **kwargs):
        if not self.has_organization_access():
            return JsonResponse({"error": "Organization required"}, status=403)

        organization = self.get_user_organization()

        # Base queryset
        reports_qs = Report.objects.filter(
            organization=organization, is_active=True
        )

        # Prepare filter data
        filter_data = {
            "start_date": request.GET.get("start_date"),
            "end_date": request.GET.get("end_date"),
            "machine": request.GET.get("machine_id"),
            "condition": request.GET.get("condition"),
            "status": request.GET.get("status"),
        }

        # Remove None values
        filter_data = {k: v for k, v in filter_data.items() if v}

        # Apply filters using ReportFilter
        filterset = ReportFilter(filter_data, queryset=reports_qs)
        reports_qs = filterset.qs

        # Apply limit for performance and ordering
        reports_qs = reports_qs.order_by("-sample_date", "-created")[:10000]

        # Create Excel workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Dashboard Reports - {organization.name}"

        # Define headers
        headers = [
            _("Lab Number"),
            _("Machine"),
            _("Component"),
            _("Lubricant"),
            _("Lubricant Hours"),
            _("Lubricant Kms"),
            _("Serial Number Code"),
            _("Sample Date"),
            _("Reception Date"),
            _("Status"),
            _("Condition"),
            _("PER Number"),
            _("Notes"),
        ]

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = str(header)  # Convert translated text to string
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row, report in enumerate(reports_qs, 2):
            worksheet.cell(row=row, column=1).value = report.lab_number
            worksheet.cell(row=row, column=2).value = (
                report.machine.name if report.machine else ""
            )
            worksheet.cell(row=row, column=3).value = (
                report.component.type.name if report.component else ""
            )
            worksheet.cell(row=row, column=4).value = report.lubricant
            worksheet.cell(row=row, column=5).value = (
                report.lubricant_hours if report.lubricant_hours else ""
            )
            worksheet.cell(row=row, column=6).value = (
                report.lubricant_kms if report.lubricant_kms else ""
            )
            worksheet.cell(row=row, column=7).value = report.serial_number_code
            worksheet.cell(row=row, column=8).value = (
                report.sample_date.strftime("%Y-%m-%d")
                if report.sample_date
                else ""
            )
            worksheet.cell(row=row, column=9).value = (
                report.reception_date.strftime("%Y-%m-%d")
                if report.reception_date
                else ""
            )
            worksheet.cell(
                row=row, column=10
            ).value = report.get_status_display()
            worksheet.cell(
                row=row, column=11
            ).value = report.get_condition_display()
            worksheet.cell(row=row, column=12).value = report.per_number
            worksheet.cell(row=row, column=13).value = report.notes

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Max width 50
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Prepare response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        filename = f"dashboard_reports_{organization.name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        workbook.save(response)
        return response


# ============================================================================
# ORGANIZATION DASHBOARD API VIEWS
# ============================================================================


class OrganizationDashboardOverviewAPIView(
    LoginRequiredMixin, OrganizationRequiredMixin, View
):
    """
    API endpoint for organization dashboard overview.

    Returns fleet summary KPIs including:
    - Total machines count
    - Critical alerts count
    - Average fleet health score
    - Reports count
    """

    def get(self, request, *args, **kwargs):
        """
        Get organization dashboard overview data.

        Returns:
            JsonResponse: Dictionary containing KPIs and summary data
        """
        if not self.has_organization_access():
            return JsonResponse({"error": "Organization required"}, status=403)

        organization = self.get_user_organization()
        filter_end_date = timezone.now()
        filter_start_date = filter_end_date.replace(day=1)

        if request.GET.get("start_date"):
            try:
                filter_start_date = timezone.datetime.strptime(
                    request.GET.get("start_date"), "%Y-%m-%d"
                )
                filter_start_date = timezone.make_aware(filter_start_date)
            except ValueError:
                pass

        if request.GET.get("end_date"):
            try:
                filter_end_date = timezone.datetime.strptime(
                    request.GET.get("end_date"), "%Y-%m-%d"
                )
                filter_end_date = timezone.make_aware(filter_end_date)
            except ValueError:
                pass

        # Get all active reports for organization
        reports_qs = Report.objects.filter(
            organization=organization,
            is_active=True,
            sample_date__gte=filter_start_date,
            sample_date__lte=filter_end_date,
        ).select_related("machine", "component", "component__type")

        # Reports this month
        reports_this_month = reports_qs.count()

        # Total unique machines
        total_machines = Machine.objects.filter(
            id__in=reports_qs.values_list("machine_id", flat=True).distinct()
        ).count()

        # Critical and caution alerts (based on condition)
        critical_alerts = reports_qs.filter(
            condition=ReportCondition.CRITICAL
        ).count()
        caution_alerts = reports_qs.filter(
            condition=ReportCondition.CAUTION
        ).count()
        total_alerts = critical_alerts + caution_alerts

        # Calculate average fleet health score (0-100)
        # Based on condition distribution
        total_reports = reports_qs.count()
        if total_reports > 0:
            normal_count = reports_qs.filter(
                condition=ReportCondition.NORMAL
            ).count()
            caution_count = reports_qs.filter(
                condition=ReportCondition.CAUTION
            ).count()
            critical_count = reports_qs.filter(
                condition=ReportCondition.CRITICAL
            ).count()

            # Score calculation: Normal=100, Caution=50, Critical=0
            avg_health_score = int(
                (
                    (normal_count * 100)
                    + (caution_count * 50)
                    + (critical_count * 0)
                )
                / total_reports
            )
        else:
            avg_health_score = 0

        # Get recent critical/caution reports (last 10)
        recent_alerts = (
            reports_qs.filter(
                Q(condition=ReportCondition.CRITICAL)
                | Q(condition=ReportCondition.CAUTION)
            )
            .order_by("-sample_date")[:10]
            .values(
                "id",
                "lab_number",
                "machine__name",
                "condition",
                "sample_date",
                "component__type__name",
            )
        )

        # Format recent alerts
        recent_alerts_data = [
            {
                "id": alert["id"],
                "lab_number": alert["lab_number"],
                "machine_name": alert["machine__name"] or "N/A",
                "component": alert["component__type__name"] or "N/A",
                "condition": dict(ReportCondition.choices).get(
                    alert["condition"], "Unknown"
                ),
                "condition_class": self._get_condition_class(
                    alert["condition"]
                ),
                "sample_date": alert["sample_date"].strftime("%Y-%m-%d")
                if alert["sample_date"]
                else "N/A",
            }
            for alert in recent_alerts
        ]

        return JsonResponse(
            {
                "total_machines": total_machines,
                "critical_alerts": critical_alerts,
                "caution_alerts": caution_alerts,
                "total_alerts": total_alerts,
                "reports_this_month": reports_this_month,
                "avg_health_score": avg_health_score,
                "recent_alerts": recent_alerts_data,
                "total_reports": total_reports,
                "filter_start_date": filter_start_date.strftime("%Y-%m-%d"),
                "filter_end_date": filter_end_date.strftime("%Y-%m-%d"),
            }
        )

    def _get_condition_class(self, condition: str) -> str:
        """Get CSS class for condition display."""
        condition_classes = {
            ReportCondition.NORMAL: "success",
            ReportCondition.CAUTION: "warning",
            ReportCondition.CRITICAL: "danger",
        }
        return condition_classes.get(condition, "secondary")
